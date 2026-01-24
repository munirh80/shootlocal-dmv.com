from fastapi import FastAPI, APIRouter, Query, HTTPException, Header, Depends, UploadFile, File
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
import math
import hashlib
import secrets
import aiofiles

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Create uploads directory
UPLOADS_DIR = ROOT_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Admin authentication
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'dmvgunrange2024')  # Default password
admin_tokens = set()  # In-memory token storage

def generate_token():
    return secrets.token_hex(32)

def verify_token(authorization: Optional[str] = Header(None)):
    if not authorization or not authorization.startswith('Bearer '):
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = authorization.replace('Bearer ', '')
    if token not in admin_tokens:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    return token

# Create the main app without a prefix
app = FastAPI(title="DMV Gun Range API", version="1.0.0")

# Mount static files for uploads
app.mount("/uploads", StaticFiles(directory=str(UPLOADS_DIR)), name="uploads")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Models
class Location(BaseModel):
    address: str
    city: str
    state: str
    zip_code: str
    latitude: Optional[float] = None
    longitude: Optional[float] = None

class Hours(BaseModel):
    monday: Optional[str] = None
    tuesday: Optional[str] = None
    wednesday: Optional[str] = None
    thursday: Optional[str] = None
    friday: Optional[str] = None
    saturday: Optional[str] = None
    sunday: Optional[str] = None

class Amenities(BaseModel):
    # Range Types
    indoor: bool = False
    outdoor: bool = False
    
    # Range Distances
    pistol_50ft: bool = False
    pistol_75ft: bool = False
    pistol_25yd: bool = False
    pistol_50yd: bool = False
    rifle_smallbore: bool = False
    rifle_centerfire: bool = False
    rifle_100yd: bool = False
    rifle_200yd: bool = False
    rifle_300yd: bool = False
    rifle_500yd: bool = False
    
    # Ammunition Types
    handgun: bool = False
    rifle: bool = False
    shotgun: bool = False
    archery: bool = False
    airgun: bool = False
    muzzle_loader: bool = False
    
    # Services
    equipment_rentals: bool = False
    instruction: bool = False
    retail_store: bool = False
    hunter_education: bool = False
    youth_programs: bool = False
    womens_programs: bool = False
    food_service: bool = False
    clubhouse: bool = False
    picnic_area: bool = False
    rv_sites: bool = False
    lodging: bool = False
    
    # Competition Types
    precision_pistol: bool = False
    practical_pistol: bool = False
    smallbore_competition: bool = False
    centerfire_competition: bool = False
    airgun_competition: bool = False
    rimfire_challenge: bool = False
    three_gun: bool = False
    cowboy_action: bool = False
    uspsa: bool = False
    idpa: bool = False
    
    # Shotgun Specific
    trap: bool = False
    skeet: bool = False
    sporting_clays: bool = False
    five_stand: bool = False
    
    # Safety & Access
    ada_accessible: bool = False
    climate_controlled: bool = False
    bulletproof_barriers: bool = False
    public_access: bool = True
    members_only: bool = False
    
    # Training
    concealed_carry_classes: bool = False
    basic_firearm_training: bool = False
    advanced_training: bool = False

class Pricing(BaseModel):
    day_pass: Optional[float] = None
    monthly_membership: Optional[float] = None
    annual_membership: Optional[float] = None
    hourly_rate: Optional[float] = None
    notes: Optional[str] = None

class Range(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    phone: Optional[str] = None
    website: Optional[str] = None
    email: Optional[str] = None
    
    location: Location
    hours: Optional[Hours] = None
    amenities: Amenities
    pricing: Optional[Pricing] = None
    
    # Additional Info
    photos: List[str] = Field(default_factory=list)
    google_rating: Optional[float] = None
    google_reviews: Optional[int] = None
    google_maps_url: Optional[str] = None
    
    # Meta
    nssf_member: bool = False
    verified: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RangeCreate(BaseModel):
    name: str
    description: Optional[str] = None
    phone: Optional[str] = None
    website: Optional[str] = None
    email: Optional[str] = None
    location: Location
    hours: Optional[Hours] = None
    amenities: Amenities
    pricing: Optional[Pricing] = None
    photos: List[str] = Field(default_factory=list)
    google_rating: Optional[float] = None
    google_reviews: Optional[int] = None
    google_maps_url: Optional[str] = None
    nssf_member: bool = False
    verified: bool = False

class SearchFilters(BaseModel):
    # Location
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    radius_miles: Optional[int] = 20
    city: Optional[str] = None
    state: Optional[str] = None
    zip_code: Optional[str] = None
    
    # Basic Filters
    indoor: Optional[bool] = None
    outdoor: Optional[bool] = None
    nssf_member: Optional[bool] = None
    public_access: Optional[bool] = None
    
    # Services
    instruction: Optional[bool] = None
    equipment_rentals: Optional[bool] = None
    retail_store: Optional[bool] = None
    youth_programs: Optional[bool] = None
    womens_programs: Optional[bool] = None
    
    # Competition Types
    uspsa: Optional[bool] = None
    idpa: Optional[bool] = None
    precision_pistol: Optional[bool] = None
    three_gun: Optional[bool] = None
    
    # Range Types
    handgun: Optional[bool] = None
    rifle: Optional[bool] = None
    shotgun: Optional[bool] = None
    archery: Optional[bool] = None

# Utility function to calculate distance between two points
def calculate_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Calculate the great circle distance between two points in miles"""
    R = 3959  # Earth's radius in miles
    
    lat1_rad = math.radians(lat1)
    lon1_rad = math.radians(lon1)
    lat2_rad = math.radians(lat2)
    lon2_rad = math.radians(lon2)
    
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad
    
    a = math.sin(dlat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    
    return R * c

# API Routes
@api_router.get("/")
async def root():
    return {"message": "DMV Gun Range API - Find shooting ranges in DC, Maryland and Virginia"}

@api_router.post("/ranges", response_model=Range)
async def create_range(range_data: RangeCreate):
    range_dict = range_data.model_dump()
    range_obj = Range(**range_dict)
    
    # Convert to dict for MongoDB
    doc = range_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    await db.ranges.insert_one(doc)
    return range_obj

@api_router.get("/ranges", response_model=List[Range])
async def search_ranges(
    # Location parameters
    latitude: Optional[float] = Query(None, description="User's latitude"),
    longitude: Optional[float] = Query(None, description="User's longitude"),
    radius: Optional[int] = Query(20, description="Search radius in miles"),
    city: Optional[str] = Query(None, description="City name"),
    state: Optional[str] = Query(None, description="State abbreviation"),
    zip_code: Optional[str] = Query(None, description="ZIP code"),
    
    # Basic filters
    indoor: Optional[bool] = Query(None, description="Indoor ranges only"),
    outdoor: Optional[bool] = Query(None, description="Outdoor ranges only"),
    nssf_member: Optional[bool] = Query(None, description="NSSF member ranges only"),
    public_access: Optional[bool] = Query(None, description="Public access ranges"),
    
    # Services
    instruction: Optional[bool] = Query(None, description="Instruction available"),
    equipment_rentals: Optional[bool] = Query(None, description="Equipment rentals"),
    retail_store: Optional[bool] = Query(None, description="Retail store on-site"),
    youth_programs: Optional[bool] = Query(None, description="Youth programs"),
    womens_programs: Optional[bool] = Query(None, description="Women's programs"),
    
    # Competition types
    uspsa: Optional[bool] = Query(None, description="USPSA matches"),
    idpa: Optional[bool] = Query(None, description="IDPA matches"),
    precision_pistol: Optional[bool] = Query(None, description="Precision pistol"),
    three_gun: Optional[bool] = Query(None, description="3-Gun competitions"),
    
    # Range types
    handgun: Optional[bool] = Query(None, description="Handgun ranges"),
    rifle: Optional[bool] = Query(None, description="Rifle ranges"),
    shotgun: Optional[bool] = Query(None, description="Shotgun ranges"),
    archery: Optional[bool] = Query(None, description="Archery ranges"),
    
    limit: Optional[int] = Query(50, description="Maximum number of results")
):
    # Build MongoDB query
    query = {}
    
    # Location-based filters
    if city:
        query["location.city"] = {"$regex": city, "$options": "i"}
    if state:
        query["location.state"] = state.upper()
    if zip_code:
        query["location.zip_code"] = zip_code
    
    # Basic filters
    if indoor is not None:
        query["amenities.indoor"] = indoor
    if outdoor is not None:
        query["amenities.outdoor"] = outdoor
    if nssf_member is not None:
        query["nssf_member"] = nssf_member
    if public_access is not None:
        query["amenities.public_access"] = public_access
    
    # Service filters
    if instruction is not None:
        query["amenities.instruction"] = instruction
    if equipment_rentals is not None:
        query["amenities.equipment_rentals"] = equipment_rentals
    if retail_store is not None:
        query["amenities.retail_store"] = retail_store
    if youth_programs is not None:
        query["amenities.youth_programs"] = youth_programs
    if womens_programs is not None:
        query["amenities.womens_programs"] = womens_programs
    
    # Competition filters
    if uspsa is not None:
        query["amenities.uspsa"] = uspsa
    if idpa is not None:
        query["amenities.idpa"] = idpa
    if precision_pistol is not None:
        query["amenities.precision_pistol"] = precision_pistol
    if three_gun is not None:
        query["amenities.three_gun"] = three_gun
    
    # Range type filters
    if handgun is not None:
        query["amenities.handgun"] = handgun
    if rifle is not None:
        query["amenities.rifle"] = rifle
    if shotgun is not None:
        query["amenities.shotgun"] = shotgun
    if archery is not None:
        query["amenities.archery"] = archery
    
    # Execute query
    ranges_cursor = db.ranges.find(query, {"_id": 0}).limit(limit)
    ranges = await ranges_cursor.to_list(length=limit)
    
    # Convert timestamps back to datetime objects
    for range_doc in ranges:
        if isinstance(range_doc.get('created_at'), str):
            range_doc['created_at'] = datetime.fromisoformat(range_doc['created_at'])
        if isinstance(range_doc.get('updated_at'), str):
            range_doc['updated_at'] = datetime.fromisoformat(range_doc['updated_at'])
    
    # Filter by distance if latitude and longitude provided
    if latitude is not None and longitude is not None:
        filtered_ranges = []
        for range_doc in ranges:
            range_lat = range_doc.get('location', {}).get('latitude')
            range_lon = range_doc.get('location', {}).get('longitude')
            
            if range_lat is not None and range_lon is not None:
                distance = calculate_distance(latitude, longitude, range_lat, range_lon)
                if distance <= radius:
                    range_doc['distance'] = distance
                    filtered_ranges.append(range_doc)
        
        # Sort by distance
        filtered_ranges.sort(key=lambda x: x.get('distance', float('inf')))
        ranges = filtered_ranges
    
    return [Range(**range_doc) for range_doc in ranges]

@api_router.get("/ranges/{range_id}", response_model=Range)
async def get_range(range_id: str):
    range_doc = await db.ranges.find_one({"id": range_id}, {"_id": 0})
    if not range_doc:
        raise HTTPException(status_code=404, detail="Range not found")
    
    # Convert timestamps
    if isinstance(range_doc.get('created_at'), str):
        range_doc['created_at'] = datetime.fromisoformat(range_doc['created_at'])
    if isinstance(range_doc.get('updated_at'), str):
        range_doc['updated_at'] = datetime.fromisoformat(range_doc['updated_at'])
    
    return Range(**range_doc)

@api_router.get("/states")
async def get_available_states():
    """Get list of available states"""
    return [{"code": "VA", "name": "Virginia"}, {"code": "MD", "name": "Maryland"}, {"code": "DC", "name": "District of Columbia"}]

# Range Submission Model
class RangeSubmission(BaseModel):
    name: str
    phone: Optional[str] = None
    website: Optional[str] = None
    email: Optional[str] = None
    address: str
    city: str
    state: str
    zip_code: str
    description: Optional[str] = None
    hours: Optional[Dict[str, str]] = None
    amenities: Optional[Dict[str, bool]] = None

@api_router.post("/ranges/submit")
async def submit_range(submission: RangeSubmission):
    """Submit a new range for review (pending approval)"""
    
    # Create the range document
    range_doc = {
        "id": str(uuid.uuid4()),
        "name": submission.name,
        "description": submission.description,
        "phone": submission.phone,
        "website": submission.website,
        "email": submission.email,
        "location": {
            "address": submission.address,
            "city": submission.city,
            "state": submission.state.upper(),
            "zip_code": submission.zip_code,
            "latitude": None,
            "longitude": None
        },
        "hours": submission.hours,
        "amenities": submission.amenities or {
            "indoor": False,
            "outdoor": False,
            "handgun": False,
            "rifle": False,
            "shotgun": False,
            "archery": False,
            "equipment_rentals": False,
            "instruction": False,
            "retail_store": False,
            "concealed_carry_classes": False,
            "basic_firearm_training": False,
            "advanced_training": False,
            "trap": False,
            "skeet": False,
            "sporting_clays": False,
            "ada_accessible": False,
            "climate_controlled": False,
            "public_access": True,
            "members_only": False
        },
        "pricing": None,
        "photos": [],
        "google_rating": None,
        "google_reviews": None,
        "google_maps_url": None,
        "nssf_member": False,
        "verified": False,  # Not verified until admin reviews
        "pending_review": True,  # Flag for admin review
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Insert into submissions collection for review
    await db.range_submissions.insert_one(range_doc)
    
    return {"message": "Range submitted successfully. It will be reviewed by our team.", "id": range_doc["id"]}

# Admin Login Model
class AdminLoginRequest(BaseModel):
    password: str

class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str

# Admin Endpoints
@api_router.post("/admin/login")
async def admin_login(request: AdminLoginRequest):
    """Authenticate admin user"""
    if request.password == ADMIN_PASSWORD:
        token = generate_token()
        admin_tokens.add(token)
        return {"success": True, "token": token}
    raise HTTPException(status_code=401, detail="Invalid password")

@api_router.post("/admin/logout")
async def admin_logout(token: str = Depends(verify_token)):
    """Logout admin user"""
    admin_tokens.discard(token)
    return {"success": True}

@api_router.post("/admin/change-password")
async def change_admin_password(request: ChangePasswordRequest, token: str = Depends(verify_token)):
    """Change admin password"""
    global ADMIN_PASSWORD
    
    # Verify current password
    if request.current_password != ADMIN_PASSWORD:
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    # Validate new password
    if len(request.new_password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters")
    
    # Update password in memory
    ADMIN_PASSWORD = request.new_password
    
    # Also update the .env file
    try:
        env_path = ROOT_DIR / '.env'
        with open(env_path, 'r') as f:
            lines = f.readlines()
        
        with open(env_path, 'w') as f:
            for line in lines:
                if line.startswith('ADMIN_PASSWORD='):
                    f.write(f'ADMIN_PASSWORD="{request.new_password}"\n')
                else:
                    f.write(line)
    except Exception as e:
        logging.error(f"Failed to update .env file: {e}")
    
    return {"success": True, "message": "Password changed successfully"}

@api_router.get("/admin/submissions")
async def get_pending_submissions(token: str = Depends(verify_token)):
    """Get all pending range submissions for admin review"""
    submissions = await db.range_submissions.find(
        {"pending_review": True},
        {"_id": 0}
    ).to_list(length=100)
    return submissions

@api_router.post("/admin/submissions/{submission_id}/approve")
async def approve_submission(submission_id: str, token: str = Depends(verify_token)):
    """Approve a range submission and add it to the main directory"""
    # Find the submission
    submission = await db.range_submissions.find_one(
        {"id": submission_id},
        {"_id": 0}
    )
    
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    
    # Remove the pending_review flag and mark as verified
    submission["pending_review"] = False
    submission["verified"] = True
    submission["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # Add to main ranges collection
    await db.ranges.insert_one(submission)
    
    # Remove from submissions collection
    await db.range_submissions.delete_one({"id": submission_id})
    
    return {"message": "Range approved and added to directory", "id": submission_id}

@api_router.post("/admin/submissions/{submission_id}/reject")
async def reject_submission(submission_id: str, token: str = Depends(verify_token)):
    """Reject a range submission"""
    result = await db.range_submissions.delete_one({"id": submission_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Submission not found")
    
    return {"message": "Submission rejected", "id": submission_id}

# Photo Upload Endpoint
@api_router.post("/upload/photo")
async def upload_photo(file: UploadFile = File(...)):
    """Upload a photo for a range"""
    # Validate file type
    allowed_types = ["image/jpeg", "image/png", "image/webp", "image/gif"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Invalid file type. Allowed: JPEG, PNG, WebP, GIF")
    
    # Generate unique filename
    file_ext = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    unique_filename = f"{uuid.uuid4()}.{file_ext}"
    file_path = UPLOADS_DIR / unique_filename
    
    # Save file
    try:
        async with aiofiles.open(file_path, 'wb') as f:
            content = await file.read()
            # Limit file size to 5MB
            if len(content) > 5 * 1024 * 1024:
                raise HTTPException(status_code=400, detail="File too large. Max size: 5MB")
            await f.write(content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save file: {str(e)}")
    
    # Return the URL path
    return {
        "success": True,
        "filename": unique_filename,
        "url": f"/uploads/{unique_filename}"
    }

@api_router.post("/ranges/{range_id}/photos")
async def add_photo_to_range(range_id: str, photo_url: str = Query(...)):
    """Add a photo URL to a range"""
    result = await db.ranges.update_one(
        {"id": range_id},
        {"$push": {"photos": photo_url}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Range not found")
    
    return {"success": True, "message": "Photo added to range"}

# ==================== REVIEWS ====================

class ReviewCreate(BaseModel):
    range_id: str
    reviewer_name: str
    rating: int = Field(..., ge=1, le=5)
    comment: str
    
class ReviewResponse(BaseModel):
    id: str
    range_id: str
    reviewer_name: str
    rating: int
    comment: str
    created_at: str
    helpful_count: int = 0

@api_router.post("/reviews")
async def create_review(review: ReviewCreate):
    """Submit a review for a range"""
    # Verify range exists
    range_doc = await db.ranges.find_one({"id": review.range_id})
    if not range_doc:
        raise HTTPException(status_code=404, detail="Range not found")
    
    review_doc = {
        "id": str(uuid.uuid4()),
        "range_id": review.range_id,
        "reviewer_name": review.reviewer_name,
        "rating": review.rating,
        "comment": review.comment,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "helpful_count": 0
    }
    
    await db.reviews.insert_one(review_doc)
    
    # Update range's average rating
    await update_range_rating(review.range_id)
    
    return {"success": True, "review_id": review_doc["id"]}

@api_router.get("/reviews/{range_id}")
async def get_reviews(range_id: str, limit: int = 20, skip: int = 0):
    """Get reviews for a specific range"""
    reviews = await db.reviews.find(
        {"range_id": range_id},
        {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(length=limit)
    
    total = await db.reviews.count_documents({"range_id": range_id})
    
    # Calculate stats
    pipeline = [
        {"$match": {"range_id": range_id}},
        {"$group": {
            "_id": None,
            "average_rating": {"$avg": "$rating"},
            "total_reviews": {"$sum": 1},
            "rating_5": {"$sum": {"$cond": [{"$eq": ["$rating", 5]}, 1, 0]}},
            "rating_4": {"$sum": {"$cond": [{"$eq": ["$rating", 4]}, 1, 0]}},
            "rating_3": {"$sum": {"$cond": [{"$eq": ["$rating", 3]}, 1, 0]}},
            "rating_2": {"$sum": {"$cond": [{"$eq": ["$rating", 2]}, 1, 0]}},
            "rating_1": {"$sum": {"$cond": [{"$eq": ["$rating", 1]}, 1, 0]}}
        }}
    ]
    
    stats_result = await db.reviews.aggregate(pipeline).to_list(length=1)
    stats = stats_result[0] if stats_result else {
        "average_rating": 0,
        "total_reviews": 0,
        "rating_5": 0, "rating_4": 0, "rating_3": 0, "rating_2": 0, "rating_1": 0
    }
    
    return {
        "reviews": reviews,
        "total": total,
        "stats": {
            "average_rating": round(stats.get("average_rating", 0), 1),
            "total_reviews": stats.get("total_reviews", 0),
            "distribution": {
                "5": stats.get("rating_5", 0),
                "4": stats.get("rating_4", 0),
                "3": stats.get("rating_3", 0),
                "2": stats.get("rating_2", 0),
                "1": stats.get("rating_1", 0)
            }
        }
    }

@api_router.post("/reviews/{review_id}/helpful")
async def mark_review_helpful(review_id: str):
    """Mark a review as helpful"""
    result = await db.reviews.update_one(
        {"id": review_id},
        {"$inc": {"helpful_count": 1}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Review not found")
    return {"success": True}

async def update_range_rating(range_id: str):
    """Update the average rating for a range"""
    pipeline = [
        {"$match": {"range_id": range_id}},
        {"$group": {"_id": None, "avg": {"$avg": "$rating"}, "count": {"$sum": 1}}}
    ]
    result = await db.reviews.aggregate(pipeline).to_list(length=1)
    
    if result:
        avg_rating = round(result[0]["avg"], 1)
        review_count = result[0]["count"]
        await db.ranges.update_one(
            {"id": range_id},
            {"$set": {"user_rating": avg_rating, "user_reviews_count": review_count}}
        )

# ==================== BULK IMPORT ====================

@api_router.post("/admin/bulk-import")
async def bulk_import_ranges(file: UploadFile = File(...), token: str = Depends(verify_token)):
    """Bulk import ranges from CSV or Excel file"""
    import csv
    import io
    
    # Validate file type
    allowed_types = ["text/csv", "application/vnd.ms-excel", 
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"]
    
    if file.content_type not in allowed_types and not file.filename.endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file type. Allowed: CSV, Excel")
    
    content = await file.read()
    
    # Parse CSV
    if file.filename.endswith('.csv') or file.content_type == "text/csv":
        try:
            decoded = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse CSV: {str(e)}")
    else:
        # For Excel files, we need openpyxl
        try:
            import openpyxl
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(content))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    rows.append(dict(zip(headers, row)))
        except ImportError:
            raise HTTPException(status_code=400, detail="Excel support requires openpyxl. Please use CSV format.")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {str(e)}")
    
    # Process and insert ranges
    imported = 0
    errors = []
    
    for idx, row in enumerate(rows):
        try:
            # Map CSV columns to our schema
            name = row.get('name') or row.get('Name') or row.get('range_name')
            if not name:
                errors.append(f"Row {idx + 2}: Missing name")
                continue
            
            range_doc = {
                "id": str(uuid.uuid4()),
                "name": name,
                "description": row.get('description') or row.get('Description'),
                "phone": row.get('phone') or row.get('Phone'),
                "website": row.get('website') or row.get('Website') or row.get('site'),
                "email": row.get('email') or row.get('Email'),
                "location": {
                    "address": row.get('address') or row.get('Address') or row.get('street') or "",
                    "city": row.get('city') or row.get('City') or "",
                    "state": (row.get('state') or row.get('State') or "").upper()[:2],
                    "zip_code": str(row.get('zip_code') or row.get('zip') or row.get('postal_code') or ""),
                    "latitude": float(row.get('latitude') or row.get('lat') or 0) if row.get('latitude') or row.get('lat') else None,
                    "longitude": float(row.get('longitude') or row.get('lon') or row.get('lng') or 0) if row.get('longitude') or row.get('lon') or row.get('lng') else None
                },
                "hours": None,
                "amenities": {
                    "indoor": str(row.get('indoor', '')).lower() in ['true', '1', 'yes', 'x'],
                    "outdoor": str(row.get('outdoor', '')).lower() in ['true', '1', 'yes', 'x'],
                    "handgun": True,
                    "rifle": True,
                    "shotgun": str(row.get('shotgun', '')).lower() in ['true', '1', 'yes', 'x'],
                    "equipment_rentals": str(row.get('rentals', '')).lower() in ['true', '1', 'yes', 'x'],
                    "instruction": str(row.get('instruction', '')).lower() in ['true', '1', 'yes', 'x'],
                    "public_access": True
                },
                "photos": [],
                "nssf_member": False,
                "verified": False,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            
            # Filter to DMV area only
            if range_doc["location"]["state"] not in ["VA", "MD", "DC"]:
                errors.append(f"Row {idx + 2}: State must be VA, MD, or DC")
                continue
            
            await db.ranges.insert_one(range_doc)
            imported += 1
            
        except Exception as e:
            errors.append(f"Row {idx + 2}: {str(e)}")
    
    return {
        "success": True,
        "imported": imported,
        "total_rows": len(rows),
        "errors": errors[:10] if errors else []  # Return first 10 errors
    }

@api_router.get("/stats")
async def get_stats():
    """Get basic statistics about ranges in the database"""
    total_ranges = await db.ranges.count_documents({})
    va_ranges = await db.ranges.count_documents({"location.state": "VA"})
    md_ranges = await db.ranges.count_documents({"location.state": "MD"})
    dc_ranges = await db.ranges.count_documents({"location.state": "DC"})
    indoor_ranges = await db.ranges.count_documents({"amenities.indoor": True})
    outdoor_ranges = await db.ranges.count_documents({"amenities.outdoor": True})
    nssf_members = await db.ranges.count_documents({"nssf_member": True})
    
    return {
        "total_ranges": total_ranges,
        "virginia_ranges": va_ranges,
        "maryland_ranges": md_ranges,
        "dc_ranges": dc_ranges,
        "indoor_ranges": indoor_ranges,
        "outdoor_ranges": outdoor_ranges,
        "nssf_members": nssf_members
    }

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)